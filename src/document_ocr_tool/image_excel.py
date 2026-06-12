from __future__ import annotations

import re
from pathlib import Path
from statistics import median
from typing import Any, Callable

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .layout_utils import cluster_lines, normalize_items
from .ocr_engine import RapidOcrEngine
from .table_utils import detect_simple_table


SUPPORTED_IMAGE_SUFFIXES = {
    ".bmp",
    ".jpeg",
    ".jpg",
    ".png",
    ".tif",
    ".tiff",
    ".webp",
}


def _read_image(image_path: Path) -> np.ndarray:
    data = np.fromfile(str(image_path), dtype=np.uint8)
    image = cv2.imdecode(data, cv2.IMREAD_COLOR)
    if image is None:
        raise ValueError(f"无法读取图片：{image_path}")
    return image


def _cluster_positions(indices: np.ndarray, max_gap: int = 4) -> list[int]:
    if len(indices) == 0:
        return []
    groups: list[list[int]] = [[int(indices[0])]]
    for value in indices[1:]:
        value = int(value)
        if value - groups[-1][-1] <= max_gap:
            groups[-1].append(value)
        else:
            groups.append([value])
    return [int(round(sum(group) / len(group))) for group in groups]


def _longest_run(values: np.ndarray) -> int:
    longest = 0
    current = 0
    for value in values:
        if value:
            current += 1
            longest = max(longest, current)
        else:
            current = 0
    return longest


def _filter_continuous_lines(
    mask: np.ndarray,
    lines: list[int],
    vertical: bool,
    minimum_length: int,
) -> list[int]:
    filtered = []
    limit = mask.shape[1] if vertical else mask.shape[0]
    for line in lines:
        samples = []
        for offset in range(-2, 3):
            position = line + offset
            if 0 <= position < limit:
                values = mask[:, position] > 0 if vertical else mask[position, :] > 0
                samples.append(_longest_run(values))
        if samples and max(samples) >= minimum_length:
            filtered.append(line)
    return filtered


def _detect_grid_lines(image: np.ndarray) -> tuple[list[int], list[int]]:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    binary = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        31,
        12,
    )
    height, width = binary.shape
    horizontal_kernel = cv2.getStructuringElement(
        cv2.MORPH_RECT,
        (max(18, width // 28), 1),
    )
    vertical_kernel = cv2.getStructuringElement(
        cv2.MORPH_RECT,
        (1, max(18, height // 28)),
    )
    horizontal = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel)
    vertical = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel)

    horizontal_strength = np.count_nonzero(horizontal, axis=1)
    vertical_strength = np.count_nonzero(vertical, axis=0)
    y_indices = np.where(horizontal_strength >= max(20, int(width * 0.18)))[0]
    x_indices = np.where(vertical_strength >= max(20, int(height * 0.18)))[0]
    y_lines = _cluster_positions(y_indices)
    x_lines = _cluster_positions(x_indices)
    y_lines = _filter_continuous_lines(
        horizontal,
        y_lines,
        vertical=False,
        minimum_length=max(20, int(width * 0.18)),
    )
    x_lines = _filter_continuous_lines(
        vertical,
        x_lines,
        vertical=True,
        minimum_length=max(20, int(height * 0.18)),
    )

    min_spacing = max(6, int(median([item for item in [height, width] if item]) * 0.004))
    return _deduplicate_lines(x_lines, min_spacing), _deduplicate_lines(y_lines, min_spacing)


def _deduplicate_lines(lines: list[int], min_spacing: int) -> list[int]:
    result: list[int] = []
    for line in lines:
        if not result or line - result[-1] >= min_spacing:
            result.append(line)
        else:
            result[-1] = int(round((result[-1] + line) / 2))
    return result


def _text_for_cell(
    items: list[dict[str, Any]],
    x1: int,
    y1: int,
    x2: int,
    y2: int,
) -> str:
    matches = []
    for item in items:
        center_x = (item["x1"] + item["x2"]) / 2
        center_y = (item["y1"] + item["y2"]) / 2
        if x1 <= center_x <= x2 and y1 <= center_y <= y2:
            matches.append(item)
    if not matches:
        return ""
    lines = cluster_lines(matches)
    return "\n".join(" ".join(part["text"] for part in line).strip() for line in lines).strip()


def _rows_from_grid(
    items: list[dict[str, Any]],
    x_lines: list[int],
    y_lines: list[int],
) -> list[list[str]] | None:
    if len(x_lines) < 3 or len(y_lines) < 3:
        return None
    if len(x_lines) > 80 or len(y_lines) > 300:
        return None

    rows = []
    for row_index in range(len(y_lines) - 1):
        row = []
        for col_index in range(len(x_lines) - 1):
            row.append(
                _text_for_cell(
                    items,
                    x_lines[col_index],
                    y_lines[row_index],
                    x_lines[col_index + 1],
                    y_lines[row_index + 1],
                )
            )
        rows.append(row)

    nonempty_rows = sum(1 for row in rows if any(cell for cell in row))
    nonempty_cells = sum(1 for row in rows for cell in row if cell)
    if nonempty_rows < 2 or nonempty_cells < 4:
        return None
    return rows


def _rows_from_ocr_layout(items: list[dict[str, Any]]) -> list[list[str]]:
    table = detect_simple_table(items)
    if table:
        return table

    normalized = normalize_items(items)
    lines = cluster_lines(normalized)
    rows = [[item["text"] for item in line] for line in lines if line]
    if not rows:
        raise ValueError("图片中没有识别到可写入 Excel 的文字。")
    return rows


def extract_table_rows(image_path: Path, ocr_items: list[dict[str, Any]]) -> tuple[list[list[str]], str]:
    image = _read_image(image_path)
    normalized = normalize_items(ocr_items)
    x_lines, y_lines = _detect_grid_lines(image)
    grid_rows = _rows_from_grid(normalized, x_lines, y_lines)
    if grid_rows:
        return grid_rows, "有边框表格"
    return _rows_from_ocr_layout(ocr_items), "无边框/文字对齐表格"


def _safe_sheet_title(stem: str) -> str:
    title = re.sub(r"[\[\]:*?/\\]", "_", stem).strip() or "识别结果"
    return title[:31]


def _write_workbook(rows: list[list[str]], output_path: Path, sheet_title: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_title(sheet_title)
    sheet.freeze_panes = "A2"

    thin_gray = Side(style="thin", color="B7B7B7")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    max_columns = max((len(row) for row in rows), default=1)
    for row_index, row in enumerate(rows, start=1):
        for col_index in range(1, max_columns + 1):
            value = row[col_index - 1] if col_index <= len(row) else ""
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if row_index == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill

    for col_index in range(1, max_columns + 1):
        values = [
            str(sheet.cell(row=row_index, column=col_index).value or "")
            for row_index in range(1, sheet.max_row + 1)
        ]
        width = min(45, max(10, max((len(line) for value in values for line in value.splitlines()), default=0) + 2))
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def convert_image_to_excel(
    image_path: Path,
    output_path: Path,
    language: str,
    logger,
    progress_callback: Callable[[int, int], None] | None = None,
) -> bool:
    if image_path.suffix.lower() not in SUPPORTED_IMAGE_SUFFIXES:
        raise ValueError(f"不支持的图片格式：{image_path.suffix}")

    logger.log(f"正在 OCR 识别图片：{image_path.name}")
    engine = RapidOcrEngine(language=language, logger=logger)
    ocr_items = engine.recognize(image_path)
    if progress_callback:
        progress_callback(1, 2)

    rows, table_type = extract_table_rows(image_path, ocr_items)
    logger.log(f"已识别为{table_type}，共 {len(rows)} 行。")
    _write_workbook(rows, output_path, image_path.stem)
    if progress_callback:
        progress_callback(2, 2)
    logger.log(f"图片转 Excel 完成：{output_path}")
    return True
