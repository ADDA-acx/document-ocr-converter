from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import cv2
import numpy as np
from openpyxl import load_workbook

from document_ocr_tool.image_excel import convert_image_to_excel, extract_table_rows


def make_item(text: str, x1: int, y1: int, x2: int, y2: int) -> dict:
    return {
        "text": text,
        "score": 0.99,
        "box": [[x1, y1], [x2, y1], [x2, y2], [x1, y2]],
    }


class FakeLogger:
    def log(self, message: str) -> None:
        pass


class ImageExcelTests(unittest.TestCase):
    def test_bordered_table_is_written_to_xlsx(self) -> None:
        values = [
            ["Name", "Qty", "Price"],
            ["Apple", "2", "3.50"],
            ["Orange", "5", "8.00"],
        ]
        xs = [20, 180, 340, 500]
        ys = [20, 90, 160, 230]
        image = np.full((260, 520, 3), 255, dtype=np.uint8)
        for x in xs:
            cv2.line(image, (x, ys[0]), (x, ys[-1]), (0, 0, 0), 3)
        for y in ys:
            cv2.line(image, (xs[0], y), (xs[-1], y), (0, 0, 0), 3)

        items = []
        for row_index, row in enumerate(values):
            for col_index, text in enumerate(row):
                items.append(
                    make_item(
                        text,
                        xs[col_index] + 15,
                        ys[row_index] + 15,
                        xs[col_index + 1] - 15,
                        ys[row_index + 1] - 15,
                    )
                )

        class FakeEngine:
            def __init__(self, *args, **kwargs) -> None:
                pass

            def recognize(self, image_path: Path) -> list[dict]:
                return items

        with tempfile.TemporaryDirectory() as temp_dir:
            image_path = Path(temp_dir) / "table.png"
            output_path = Path(temp_dir) / "table.xlsx"
            cv2.imencode(".png", image)[1].tofile(str(image_path))
            with patch("document_ocr_tool.image_excel.RapidOcrEngine", FakeEngine):
                convert_image_to_excel(
                    image_path,
                    output_path,
                    language="chinese_english",
                    logger=FakeLogger(),
                )

            sheet = load_workbook(output_path).active
            actual = [
                [sheet.cell(row, column).value or "" for column in range(1, 4)]
                for row in range(1, 4)
            ]
            self.assertEqual(values, actual)

    def test_borderless_table_falls_back_to_ocr_alignment(self) -> None:
        image = np.full((180, 420, 3), 255, dtype=np.uint8)
        items = [
            make_item("Product", 20, 20, 100, 45),
            make_item("Qty", 180, 20, 220, 45),
            make_item("Apple", 20, 80, 85, 105),
            make_item("2", 180, 80, 195, 105),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            image_path = Path(temp_dir) / "borderless.png"
            cv2.imencode(".png", image)[1].tofile(str(image_path))
            rows, table_type = extract_table_rows(image_path, items)

        self.assertEqual("无边框/文字对齐表格", table_type)
        self.assertEqual([["Product", "Qty"], ["Apple", "2"]], rows)


if __name__ == "__main__":
    unittest.main()
